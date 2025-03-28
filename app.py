import streamlit as st
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import requests
import os
import shutil
import tempfile
from io import BytesIO

# === PAGE CONFIG (must be first Streamlit command) ===
st.set_page_config(page_title="Excel Image Inserter", layout="centered")

# === LOGO + HEADER ===
logo_path = "assets/logo.png"
logo = Image.open(logo_path)

col1, col2 = st.columns([1, 6])
with col1:
    st.image(logo, width=100)
with col2:
    st.title("Upload scraped data")
    st.caption("This tool adds images to your Excel file based on the 'image url' column.")

# === SETTINGS ===
IMAGE_URL_HEADER = "image url"
IMAGE_WIDTH = 100
IMAGE_HEIGHT = 100

# === FILE UPLOAD ===
uploaded_file = st.file_uploader("Upload Excel file (.xlsx)", type=["xlsx"])

if uploaded_file:
    with tempfile.TemporaryDirectory() as temp_dir:
        # Save uploaded file temporarily
        file_path = os.path.join(temp_dir, "input.xlsx")
        with open(file_path, "wb") as f:
            f.write(uploaded_file.read())

        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Find the column with the image URLs
        url_col_index = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == IMAGE_URL_HEADER:
                url_col_index = col
                break

        if url_col_index is None:
            st.error(f"❌ Couldn't find a column named '{IMAGE_URL_HEADER}'. Please check your file.")
            st.stop()

        # Add a header for the image column
        image_col_index = ws.max_column + 1
        image_col_letter = openpyxl.utils.get_column_letter(image_col_index)
        ws[f"{image_col_letter}1"] = "Image"

        # Create temp folder for images
        image_temp_dir = os.path.join(temp_dir, "images")
        os.makedirs(image_temp_dir, exist_ok=True)

        # Insert images
        for row in range(2, ws.max_row + 1):
            url = ws.cell(row=row, column=url_col_index).value
            if url and str(url).startswith("http"):
                try:
                    response = requests.get(url)
                    if response.status_code == 200:
                        img_path = os.path.join(image_temp_dir, f"img_{row}.png")
                        with open(img_path, "wb") as img_file:
                            img_file.write(response.content)

                        img = XLImage(img_path)
                        img.width = IMAGE_WIDTH
                        img.height = IMAGE_HEIGHT

                        ws.row_dimensions[row].height = IMAGE_HEIGHT * 0.75
                        img_cell = f"{image_col_letter}{row}"
                        ws.add_image(img, img_cell)
                except Exception as e:
                    st.warning(f"⚠️ Error processing row {row}: {e}")

        # Save new Excel to memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Provide download button
        st.success("✅ Done! Download updated file.")
        st.download_button(
            label="📤 Download Result File",
            data=output,
            file_name="with_images.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
